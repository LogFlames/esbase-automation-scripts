from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC 
import openpyxl
import argparse
import re
import time

def update_accnr(driver, row):
    driver.get("http://esbase.nrm.se/accession?id=" + row["accnr"] + "&og")
    time.sleep(0.1)

    elem_first = None
    old_values = []
    for j in range(len(row["property_id"])):
        try:
            WebDriverWait(driver, 0.1, poll_frequency=0.05).until(EC.alert_is_present(), "No alert")
            driver.switch_to.alert.accept()
            print("Dismissed alert")
        except TimeoutException:
            pass

        try:
            elem = driver.find_element(By.ID, row["property_id"][j])
            if j == 0:
                elem_first = elem
        except NoSuchElementException:
            raise Exception(f"Invalid property_id '{row['property_id'][j]}', could not find element with id.")

        if elem.tag_name == "select":
            s = Select(elem)
            old_values.append(s.first_selected_option.text)
            s.select_by_visible_text(row["new_value"][j])
        else:
            old_values.append(elem.get_attribute("value"))
            elem.clear()
            elem.send_keys(row["new_value"][j])

    return elem_first, old_values


def main():
    parser = argparse.ArgumentParser(
            prog = "Esbase Automation Script",
            description = "Automates interactions with the php script using selenium")
    parser.add_argument("--excel", dest = "excel", help = "The path to an excel file containing three columns, with the names: 'accnr', 'property_id', 'new_value'", required = True)

    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.excel).active
    if wb.max_column % 2 != 1:
        raise Exception("The excel file must have an odd number of columns")

    to_change = []
    for row in range(0, wb.max_row):
        values = []
        for col in wb.iter_cols(1, wb.max_column):
            val = col[row].value
            if val is None:
                val = ""
            val = str(val)
            values.append(val)
        if row == 0:
            if values[0] != "accnr":
                raise Exception("The first column in the Excel-file must be accnr")
            for prop in range(1, len(values), 2):
                if "property_id" not in values[prop]:
                    raise Exception(f"The column {prop + 1} must contain 'property_id'")
            for nv in range(2, len(values), 2):
                if "new_value" not in values[nv]:
                    raise Exception(f"The column {nv + 1} must contain 'new_value'")
        else:
            if re.match("^[ABCDGHLXP][0-9]{4}/?[0-9]{5}$", values[0]) is None:
                raise Exception(f"Invalid accnr '{values[0]}', please enter on the form: 'A2022/12345' or 'A202212345'. Only ABCDGHLXP are allowed as prefixed letters.")

            letter = values[0][0]
            year = values[0][1:5]
            if "/" in values[0]:
                value = values[0][6:]
            else:
                value = values[0][5:]

            fn = {"A": "1", "B": "2", "C": "3", "D": "4", "G": "7", "H": "8", "L": "5", "P": "9", "X": "6"}[letter]
            accdb = f"{fn}{year[1:]}{value}"

            to_change.append({ "accnr": accdb, "property_id": values[1::2], "new_value": values[2::2] })
            print(to_change[-1])

    opt = webdriver.FirefoxOptions()
    opt.binary_location = "./FirefoxPortable/App/Firefox64/firefox.exe"

    driver = webdriver.Firefox(options = opt)
    driver.implicitly_wait(60)
    driver.get("http://esbase.nrm.se")

    input(" -- Press enter when you've logged in: -- \n")


    first = True
    start = time.time()

    with open("log.txt", "a") as f:
        f.write(f"Run started at {start}" + "\n")
    with open("log.csv", "a") as f:
        f.write(f"accnr,property_id,old_value,new_value" + "\n")

    for i, row in enumerate(to_change):
        elem_first, old_values = None, None
        while elem_first is None or old_values is None:
            elem_first, old_values = update_accnr(driver, row)

        if first:
            ans = input("Does everything look good? (y/n) ")
            while ans not in ["y", "n"]:
                ans = input("Does everything look good? (y/n) ")
            if ans != "y":
                print("Aborting operations...")
                with open("log.txt", "a") as f:
                    f.write("Run aborted" + "\n")
                driver.close()
                exit()
            first = False
            start = time.time()

        if elem_first is None:
            raise Exception("First element was None, cannot submit")
        else:
            elem_first.submit()
            while "&og" in driver.current_url:
                time.sleep(0.5)

        log_txt = f"Updated '{row['accnr']}': '{row['property_id']}' from '{old_values}' to '{row['new_value']}'"
        log_csv = f"{row['accnr']}," + ",".join([f"{row['property_id'][j]},'{old_values[j]}','{row['new_value'][j]}'" for j in range(len(old_values))])
        now = time.time()
        eta = int((now - start) * len(to_change) / (i + 1))
        print(log_txt + "\tPassed time/Est total\t" + f"{int(now - start) // 60}:{int(now - start) % 60:02}/{eta//60}:{eta % 60:02}\t({i + 1}/{len(to_change)})")
        with open("log.txt", "a") as f:
            f.write(log_txt + "\n")
        with open("log.csv", "a") as f:
            f.write(log_csv + "\n")

    driver.close()

if __name__ == "__main__":
    main()
